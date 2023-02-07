import random
import tkinter.messagebox
from tkinter import *
import pandas as pd
import openpyxl as op
import re


def logout(window):
    window.destroy()
    page()


def dashboardFunc(page, window, page2, page3, page4, page5, page6, page7, page8, page9, page10, number, colNum):
    entry_code = Entry(page, highlightthickness=0, relief=FLAT, bg="#36454F", fg="#6b6a69",
                       font=("yu gothic ui", 105, "bold"))

    entry_code.place(x=70, y=140, height=150, width=460)

    loginButtonOne = Button(page, text="Click to Check-in Student", font=('yu gothic ui', 25, 'bold'), fg="white",
                            bg="#36454F",
                            borderwidth=0,
                            command=lambda: crosscheckOne(entry_code, number=number, columnNumber=colNum))
    loginButtonOne.place(x=90, y=350)

    mainSetup(page=page, page2=page2, page3=page3, page4=page4, entry=entry_code, window=window, page5=page5,
              page6=page6, page7=page7, page8=page8, page9=page9, page10=page10)


def finalWinner(tot):
    df = pd.read_excel('eventAttendance.xlsx', engine='openpyxl', dtype=object, header=None)
    list1 = df.values.tolist()
    list1.pop(0)
    del list1[tot:-1:]
    mainList = []
    i = 0
    while i <= tot:
        addListExtract = list1[i]
        addListCleaned = [int(j) for j in addListExtract]
        addList = sum(addListCleaned)
        mainList.append(addList)
        i = i + 1

    leaderScore = max(mainList)
    winnerIndex = mainList.index(leaderScore)

    podiumList = mainList.copy()
    podiumList.sort()
    secondIndexExtract = podiumList[-2]
    secondIndex = mainList.index(secondIndexExtract)
    thirdIndexExtract = podiumList[-3]
    thirdIndex = mainList.index(thirdIndexExtract)

    df2 = pd.read_excel('Book1.xlsx')
    firstNameList = df2['First Name'].to_list()
    lastNameList = df2['Last Name'].to_list()

    dfNames = op.load_workbook('Book1.xlsx')
    wsNames = dfNames.active
    wsNames.cell(row=winnerIndex + 2, column=7, value='First')
    wsNames.cell(row=secondIndex + 2, column=7, value='Second')
    wsNames.cell(row=thirdIndex + 2, column=7, value='Third')
    dfNames.save('Book1.xlsx')

    print(f'{firstNameList[winnerIndex]} {lastNameList[winnerIndex]}')

    print(mainList)
    print(secondIndexExtract)
    print(firstNameList)
    print(thirdIndexExtract)

    df3 = op.load_workbook('eventAttendance.xlsx')
    ws = df3.active
    ws.cell(row=winnerIndex + 2, column=5, value='1')
    df3.save('eventAttendance.xlsx')
    tkinter.messagebox.showinfo('Winner!', f'{firstNameList[winnerIndex]} {lastNameList[winnerIndex]} is the WINNER!')
    tkinter.messagebox.showinfo('Winner!',
                                f'{firstNameList[secondIndex]} {lastNameList[secondIndex]} is the 2nd Place WINNER!')
    tkinter.messagebox.showinfo('Winner!',
                                f'{firstNameList[thirdIndex]} {lastNameList[thirdIndex]} is the 3rd Place WINNER!')


def crosscheckOne(entry1, number, columnNumber):
    entry = entry1.get()
    df = pd.read_excel('eventCodes.xlsx', sheet_name=0)
    eventOneListInt = df[f'Event {number} Code'].to_list()

    eventOneList = [str(i) for i in eventOneListInt]

    print(entry)
    print(eventOneList)
    df2 = pd.read_excel('Book1.xlsx', sheet_name=0)

    df3 = op.load_workbook('eventAttendance.xlsx')
    ws = df3.active

    firstNameList = df2['First Name'].to_list()
    lastNameList = df2['Last Name'].to_list()
    if entry in eventOneList:
        ind = eventOneList.index(entry)
        firstName = firstNameList[ind]
        lastName = lastNameList[ind]
        tkinter.messagebox.showinfo('Check In', f'{firstName} {lastName} checked in successfully for Event {number}!')
        ws.cell(row=(ind + 2), column=columnNumber, value='1')
        df3.save('eventAttendance.xlsx')

    else:
        tkinter.messagebox.showerror("Error", "Invalid Code")


def show_frame(interior, entry, window, number):
    interior.tkraise()
    window.title(f'Event {number}')
    entry.delete(0, END)


def show_frame_no_entry(interior):
    interior.tkraise()


def mainSetup(page, page2, page3, page4, page5, page6, page7, page8, page9, page10, entry, window):
    # Page 1
    instructText = Label(page, text="Enter Event Code:", font=('yu gothic ui', 35, 'bold'), bg='#36454F',
                         fg='white')
    instructText.place(x=100, y=50)

    line = Canvas(page, width=460, height=2.0, bg="#bdb9b1", highlightthickness=0)
    line.place(x=70, y=300)

    # Flip Buttons
    sportEventOneButton = Button(page, text='Sport 1', font=('yu gothic ui', 15, 'bold'),
                                 command=lambda: show_frame(interior=page, entry=entry, window=window, number='1'),
                                 bg='#36454F', fg='white')
    sportEventTwoButton = Button(page, text='Sport 2', font=('yu gothic ui', 15, 'bold'),
                                 command=lambda: show_frame(interior=page2, entry=entry, window=window, number='2'),
                                 bg='#36454F', fg='white')
    sportEventThreeButton = Button(page, text='Sport 3', font=('yu gothic ui', 15, 'bold'),
                                   command=lambda: show_frame(interior=page3, entry=entry, window=window, number='3'),
                                   bg='#36454F', fg='white')
    sportEventFourButton = Button(page, text='Sport 4', font=('yu gothic ui', 15, 'bold'),
                                  command=lambda: show_frame(interior=page4, entry=entry, window=window, number='4'),
                                  bg='#36454F', fg='white')
    sportEventFiveButton = Button(page, text='Sport 5', font=('yu gothic ui', 15, 'bold'),
                                  command=lambda: show_frame(interior=page5, entry=entry, window=window, number='5'),
                                  bg='#36454F', fg='white')

    funEventOneButton = Button(page, text='Fun 1', font=('yu gothic ui', 15, 'bold'),
                               command=lambda: show_frame(interior=page6, entry=entry, window=window, number='6'),
                               bg='#36454F', fg='white')
    funEventTwoButton = Button(page, text='Fun 2', font=('yu gothic ui', 15, 'bold'),
                               command=lambda: show_frame(interior=page7, entry=entry, window=window, number='7'),
                               bg='#36454F', fg='white')
    funEventThreeButton = Button(page, text='Fun 3', font=('yu gothic ui', 15, 'bold'),
                                 command=lambda: show_frame(interior=page8, entry=entry, window=window, number='8'),
                                 bg='#36454F', fg='white')
    funEventFourButton = Button(page, text='Fun 4', font=('yu gothic ui', 15, 'bold'),
                                command=lambda: show_frame(interior=page9, entry=entry, window=window, number='9'),
                                bg='#36454F', fg='white')
    funEventFiveButton = Button(page, text='Fun 5', font=('yu gothic ui', 15, 'bold'), bg='#36454F', fg='white',
                                command=lambda: show_frame(interior=page10, entry=entry, window=window, number='10'))

    # Max Columns
    df = pd.read_excel('Book1.xlsx')
    maxCol = len(df['First Name'])

    winnerButton = Button(page, text='Final Winner', font=('yu gothic ui', 15, 'bold'), bg='#36454F', fg='white',
                          command=lambda: finalWinner(tot=maxCol))

    logoutButton = Button(page, text='Logout', font=('yu gothic ui', 15, 'bold'), bg='#36454F', fg='white',
                          command=lambda: logout(window))

    sportEventOneButton.place(x=0, y=450, width=120)
    sportEventTwoButton.place(x=120, y=450, width=120)
    sportEventThreeButton.place(x=240, y=450, width=120)
    sportEventFourButton.place(x=360, y=450, width=120)
    sportEventFiveButton.place(x=480, y=450, width=120)

    funEventOneButton.place(x=0, y=500, width=120)
    funEventTwoButton.place(x=120, y=500, width=120)
    funEventThreeButton.place(x=240, y=500, width=120)
    funEventFourButton.place(x=360, y=500, width=120)
    funEventFiveButton.place(x=480, y=500, width=120)

    winnerButton.place(x=0, y=550, width=480)
    logoutButton.place(x=480, y=550, width=120)


def MainEventSupervisor():
    window = Tk()
    window.rowconfigure(0, weight=1)
    window.columnconfigure(0, weight=1)
    window.geometry('600x600')

    page1 = Frame(window, bg='#36454F')
    page2 = Frame(window, bg='#36454F')
    page3 = Frame(window, bg='#36454F')
    page4 = Frame(window, bg='#36454F')
    page5 = Frame(window, bg='#36454F')
    page6 = Frame(window, bg='#36454F')
    page7 = Frame(window, bg='#36454F')
    page8 = Frame(window, bg='#36454F')
    page9 = Frame(window, bg='#36454F')
    page10 = Frame(window, bg='#36454F')

    for frame in (page1, page2, page3, page4, page5, page6, page7, page8, page9, page10):
        frame.grid(row=0, column=0, sticky='nsew')

    show_frame_no_entry(page1)

    # Page 1
    dashboardFunc(page=page1, page2=page2, page3=page3, page4=page4, page5=page5, page6=page6, page7=page7, page8=page8,
                  page9=page9, page10=page10, window=window, number='One', colNum=1)
    # Page 2
    dashboardFunc(page=page2, page2=page1, page3=page3, page4=page4, page5=page5, page6=page6, page7=page7, page8=page8,
                  page9=page9, page10=page10, window=window, number='Two', colNum=2)
    # Page 3
    dashboardFunc(page=page3, page2=page1, page3=page2, page4=page4, page5=page5, page6=page6, page7=page7, page8=page8,
                  page9=page9, page10=page10, window=window, number='Three', colNum=3)
    # Page 4
    dashboardFunc(page=page4, page2=page1, page3=page2, page4=page3, page5=page5, page6=page6, page7=page7, page8=page8,
                  page9=page9, page10=page10, window=window, number='Four', colNum=4)
    # Page 5
    dashboardFunc(page=page5, page2=page1, page3=page2, page4=page3, page5=page4, page6=page6, page7=page7, page8=page8,
                  page9=page9, page10=page10, window=window, number='Five', colNum=5)
    # Page 6
    dashboardFunc(page=page6, page2=page1, page3=page2, page4=page3, page5=page4, page6=page5, page7=page7, page8=page8,
                  page9=page9, page10=page10, window=window, number='Six', colNum=6)
    # Page 7
    dashboardFunc(page=page7, page2=page1, page3=page2, page4=page3, page5=page4, page6=page5, page7=page6, page8=page8,
                  page9=page9, page10=page10, window=window, number='Seven', colNum=7)
    # Page 8
    dashboardFunc(page=page8, page2=page1, page3=page2, page4=page3, page5=page4, page6=page5, page7=page6, page8=page7,
                  page9=page9, page10=page10, window=window, number='Eight', colNum=8)
    # Page 9
    dashboardFunc(page=page9, page2=page1, page3=page2, page4=page3, page5=page4, page6=page5, page7=page6, page8=page7,
                  page9=page8, page10=page10, window=window, number='Nine', colNum=9)
    # Page 10
    dashboardFunc(page=page10, page2=page1, page3=page2, page4=page3, page5=page4, page6=page5, page7=page6,
                  page8=page7, page9=page8, page10=page9, window=window, number='Ten', colNum=10)

    window.mainloop()


def yourPosition(z, page1, x, y, tot, x2, y2, x3, y3, x4, y4):
    df = pd.read_excel('eventAttendance.xlsx', engine='openpyxl', dtype=object, header=None)
    list1 = df.values.tolist()
    list1.pop(0)
    del list1[tot:-1:]
    nameParticipationListExtract = list1[z - 2]
    nameParticipationList = [int(i) for i in nameParticipationListExtract]
    finalScore = sum(nameParticipationList)
    showScore = Label(page1, text=finalScore, font=('yu gothic ui', 16, 'bold'), bg='#36454F', fg='white')
    showScore.place(x=x, y=y)
    if 6 > finalScore >= 3:
        showPrize = Label(page1, text="Homework Pass", font=('yu gothic ui', 16, 'bold'), bg='#36454F', fg='white')
        showPrize.place(x=x4, y=y4)
    if 9 > finalScore >= 6:
        showPrize = Label(page1, text="Pizza", font=('yu gothic ui', 16, 'bold'), bg='#36454F', fg='white')
        showPrize.place(x=x4, y=y4)
    if finalScore >= 9:
        showPrize = Label(page1, text="School T-Shirt", font=('yu gothic ui', 16, 'bold'), bg='#36454F', fg='white')
        showPrize.place(x=x4, y=y4)

    # LEADER
    mainList = []
    i = 0
    while i <= tot:
        addListExtract = list1[i]
        addListCleaned = [int(j) for j in addListExtract]
        addList = sum(addListCleaned)
        mainList.append(addList)
        i = i + 1
    leaderScore = max(mainList)
    leaderScoreIndex = mainList.index(leaderScore)
    leaderScoreLabel = Label(page1, text=leaderScore, font=('yu gothic ui', 15, 'bold'), bg='#36454F', fg='white')
    leaderScoreLabel.place(x=x2, y=y2)

    df2 = pd.read_excel('Book1.xlsx')
    firstNameList = df2['First Name'].to_list()
    lastNameList = df2['Last Name'].to_list()

    print(firstNameList)
    print(lastNameList)

    firstName = firstNameList[leaderScoreIndex]
    lastName = lastNameList[leaderScoreIndex]
    leaderLabelName = Label(page1, text=firstName + ' ' + lastName,
                            font=('yu gothic ui', 15, 'bold'), bg='#36454F', fg='white')
    leaderLabelName.place(x=x3, y=y3)


def openInfoSport(var, Type, yourcode, sport):
    newWindow = Tk()
    newWindow.geometry('600x500')
    newWindow.config(bg='#36454F')
    headerLabel = Label(newWindow, text=f'{Type} Event {var}: {sport}', font=('yu gothic ui', 40, 'bold'), bg='#36454F',
                        fg='white')
    headerLabel.place(x=20, y=20)

    infoLabel = Label(newWindow,
                      text='Lorem ipsum dolor sit\namet, consectetur adipiscing elit,\nsed do eiusmod tempor\n'
                           'incididunt ut labore et\ndolore magna aliqua.',
                      font=('yu gothic ui', 25), bg='#36454F',
                      fg='white', justify=LEFT)
    infoLabel.place(x=20, y=100)

    yourCode = Label(newWindow, text=f'Your Code: {yourcode}', font=('yu gothic ui', 35, 'bold'), bg='#36454F',
                     fg='white')
    yourCode.place(x=20, y=400)

    newWindow.mainloop()


def show_frame(interior):
    interior.tkraise()


def eventoneCode(window, eventonecode, index, number, intNumber):
    label = Label(window, text=eventonecode, font=('yu gothic ui', 65, 'bold'),
                  bg='#36454F', fg='white')
    label.place(x=270, y=200)

    wb = op.load_workbook('eventCodes.xlsx')
    ws = wb.active
    wb2 = pd.read_excel('eventCodes.xlsx', sheet_name=0)
    if eventonecode not in wb2[f'Event {number} Code']:
        ws.cell(row=index, column=intNumber, value=eventonecode)
        wb.save('eventCodes.xlsx')


def dashboard(firstName, lastName, schoolName, schoolYear, emailID, index):
    # Setup
    eventOneCode = f'{random.randint(1, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)} '
    eventTwoCode = f'{random.randint(1, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)} '
    eventThreeCode = f'{random.randint(1, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)} '
    eventFourCode = f'{random.randint(1, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)} '
    eventFiveCode = f'{random.randint(1, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)} '
    eventSixCode = f'{random.randint(1, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)} '
    eventSevenCode = f'{random.randint(1, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)} '
    eventEightCode = f'{random.randint(1, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)} '
    eventNineCode = f'{random.randint(1, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)} '
    eventTenCode = f'{random.randint(1, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)}{random.randint(0, 9)} '

    ws = pd.read_excel('eventCodes.xlsx')

    eventOneCodeList = ws['Event One Code'].to_list()
    eventTwoCodeList = ws['Event Two Code'].to_list()
    eventThreeCodeList = ws['Event Three Code'].to_list()
    eventFourCodeList = ws['Event Four Code'].to_list()
    eventFiveCodeList = ws['Event Five Code'].to_list()
    eventSixCodeList = ws['Event Six Code'].to_list()
    eventSevenCodeList = ws['Event Seven Code'].to_list()
    eventEightCodeList = ws['Event Eight Code'].to_list()
    eventNineCodeList = ws['Event Nine Code'].to_list()
    eventTenCodeList = ws['Event Ten Code'].to_list()

    print(index)
    eventOneCodeExtract = eventOneCodeList[index - 2]
    eventTwoCodeExtract = eventTwoCodeList[index - 2]
    eventThreeCodeExtract = eventThreeCodeList[index - 2]
    eventFourCodeExtract = eventFourCodeList[index - 2]
    eventFiveCodeExtract = eventFiveCodeList[index - 2]
    eventSixCodeExtract = eventSixCodeList[index - 2]
    eventSevenCodeExtract = eventSevenCodeList[index - 2]
    eventEightCodeExtract = eventEightCodeList[index - 2]
    eventNineCodeExtract = eventNineCodeList[index - 2]
    eventTenCodeExtract = eventTenCodeList[index - 2]

    # ============= Page 1 - Login =========
    window = Tk()
    window.rowconfigure(0, weight=1)
    window.columnconfigure(0, weight=1)
    window.geometry('800x600')

    page1 = Frame(window, bg='#36454F')
    page2 = Frame(window, bg='#36454F')
    page3 = Frame(window, bg='#36454F')
    page4 = Frame(window, bg='#36454F')

    for frame in (page1, page2, page3, page4):
        frame.grid(row=0, column=0, sticky='nsew')

    show_frame(page1)
    # ======== Page 2 ===========
    homeButton = Button(page1, text='Home', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page1),
                        bg='#36454F', fg='white')
    eventButton = Button(page1, text='Events', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page2),
                         bg='#36454F', fg='white')
    attendanceButton = Button(page1, text='Participation', font=('yu gothic ui', 15, 'bold'),
                              command=lambda: show_frame(page3), bg='#36454F', fg='white')
    lotteryButton = Button(page1, text='Prizes', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page4),
                           bg='#36454F', fg='white')

    # place the buttons

    homeButton.place(x=0, y=0, width=200)
    eventButton.place(x=200, y=0, width=200)
    attendanceButton.place(x=400, y=0, width=200)
    lotteryButton.place(x=600, y=0, width=200)

    # Define labels to display Personal Information
    headingLabel = Label(page1, text='Personal Information', font=('yu gothic ui', 30, 'bold'), bg='#36454F',
                         fg='white')
    nameHeaderLabel = Label(page1, text='Name: ', font=('yu gothic ui', 15, 'bold'),
                            bg='#36454F', fg='white')
    nameLabel = Label(page1, text=firstName + ' ' + lastName, font=('yu gothic ui', 15, 'bold'),
                      bg='#36454F', fg='white')
    schoolNameHeaderLabel = Label(page1, text='School Name: ', font=('yu gothic ui', 15, 'bold'),
                                  bg='#36454F', fg='white')
    schoolNameLabel = Label(page1, text=schoolName, font=('yu gothic ui', 15, 'bold'),
                            bg='#36454F', fg='white')
    schoolYearHeaderLabel = Label(page1, text='School Year: ', font=('yu gothic ui', 15, 'bold'),
                                  bg='#36454F', fg='white')
    schoolYearLabel = Label(page1, text=schoolYear, font=('yu gothic ui', 15, 'bold'),
                            bg='#36454F', fg='white')
    emailHeaderLabel = Label(page1, text='Email ID: ', font=('yu gothic ui', 15, 'bold'),
                             bg='#36454F', fg='white')
    emailLabel = Label(page1, text=emailID, font=('yu gothic ui', 15, 'bold'),
                       bg='#36454F', fg='white')

    logoutButton = Button(page1, text='Logout', font=('yu gothic ui', 15, 'bold'), bg='#36454F', fg='white',
                          command=lambda: logout(window))

    logoutButton.place(x=715, y=550)

    # Place Labels
    headingLabel.place(x=30, y=100)
    nameLabel.place(x=100, y=180)
    schoolNameLabel.place(x=165, y=240)
    schoolYearLabel.place(x=150, y=300)
    emailLabel.place(x=120, y=360)

    nameHeaderLabel.place(x=30, y=180)
    schoolNameHeaderLabel.place(x=30, y=240)
    schoolYearHeaderLabel.place(x=30, y=300)
    emailHeaderLabel.place(x=30, y=360)

    # Event Buttons
    eventOneButton = Button(page3, text='Event One', font=('yu gothic ui', 18, 'bold'), bg='#36454F',
                            fg='white',
                            command=lambda: eventoneCode(window, eventOneCode, index, number='One', intNumber=1))
    eventTwoButton = Button(page3, text='Event Two', font=('yu gothic ui', 18, 'bold'), bg='#36454F',
                            fg='white',
                            command=lambda: eventoneCode(window, eventTwoCode, index, number='Two', intNumber=2))
    eventThreeButton = Button(page3, text='Event Three', font=('yu gothic ui', 18, 'bold'), bg='#36454F',
                              fg='white',
                              command=lambda: eventoneCode(window, eventThreeCode, index, number='Three', intNumber=3))
    eventFourButton = Button(page3, text='Event Four', font=('yu gothic ui', 18, 'bold'), bg='#36454F',
                             fg='white',
                             command=lambda: eventoneCode(window, eventFourCode, index, number='Four', intNumber=4))
    eventFiveButton = Button(page3, text='Event Five', font=('yu gothic ui', 18, 'bold'), bg='#36454F',
                             fg='white',
                             command=lambda: eventoneCode(window, eventFiveCode, index, number='Five', intNumber=5))
    eventSixButton = Button(page3, text='Event Six', font=('yu gothic ui', 18, 'bold'), bg='#36454F',
                            fg='white',
                            command=lambda: eventoneCode(window, eventSixCode, index, number='Six', intNumber=6))
    eventSevenButton = Button(page3, text='Event Seven', font=('yu gothic ui', 18, 'bold'), bg='#36454F',
                              fg='white',
                              command=lambda: eventoneCode(window, eventSevenCode, index, number='Seven', intNumber=7))
    eventEightButton = Button(page3, text='Event Eight', font=('yu gothic ui', 18, 'bold'), bg='#36454F',
                              fg='white',
                              command=lambda: eventoneCode(window, eventEightCode, index, number='Eight', intNumber=8))
    eventNineButton = Button(page3, text='Event Nine', font=('yu gothic ui', 18, 'bold'), bg='#36454F',
                             fg='white',
                             command=lambda: eventoneCode(window, eventNineCode, index, number='Nine', intNumber=9))
    eventTenButton = Button(page3, text='Event Nine', font=('yu gothic ui', 18, 'bold'), bg='#36454F',
                            fg='white',
                            command=lambda: eventoneCode(window, eventTenCode, index, number='Ten', intNumber=10))

    eventOneButton.place(x=0, y=500, width=160)
    eventTwoButton.place(x=160, y=500, width=160)
    eventThreeButton.place(x=320, y=500, width=160)
    eventFourButton.place(x=480, y=500, width=160)
    eventFiveButton.place(x=640, y=500, width=160)
    eventSixButton.place(x=0, y=550, width=160)
    eventSevenButton.place(x=160, y=550, width=160)
    eventEightButton.place(x=320, y=550, width=160)
    eventNineButton.place(x=480, y=550, width=160)
    eventTenButton.place(x=640, y=550, width=160)

    infoLabel = Label(page3, text='Your Code is:', font=('yu gothic ui', 25, 'bold'), bg='#36454F', fg='white')
    infoLabel.place(x=295, y=140)

    # ======== Page 2 ===========
    homeButton = Button(page2, text='Home', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page1),
                        bg='#36454F', fg='white')
    eventButton = Button(page2, text='Events', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page2),
                         bg='#36454F', fg='white')
    attendanceButton = Button(page2, text='Participation', font=('yu gothic ui', 15, 'bold'),
                              command=lambda: show_frame(page3), bg='#36454F', fg='white')
    lotteryButton = Button(page2, text='Prizes', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page4),
                           bg='#36454F', fg='white')

    headerLabel = Label(page2, text='Upcoming Events:', font=('yu gothic ui', 30, 'bold'), bg='#36454F',
                        fg='white')
    headerLabel.place(x=30, y=100)

    eventOneDisplayButton = Button(page2, text='Sport\nOne', font=('yu gothic ui', 30, 'bold'), bg='#556d7c',
                                   fg='black',
                                   command=lambda: openInfoSport('1', 'Sport', eventOneCodeExtract, sport="Basketball"))
    eventOneDisplayButton.place(x=30, y=200, width=140, height=150)
    eventTwoDisplayButton = Button(page2, text='Sport\nTwo', font=('yu gothic ui', 30, 'bold'), bg='#556d7c',
                                   fg='black',
                                   command=lambda: openInfoSport('2', 'Sport', eventTwoCodeExtract, sport="Baseball"))
    eventTwoDisplayButton.place(x=180, y=200, width=140, height=150)
    eventThreeDisplayButton = Button(page2, text='Sport\nThree', font=('yu gothic ui', 30, 'bold'), bg='#556d7c',
                                     fg='black',
                                     command=lambda: openInfoSport('3', 'Sport', eventThreeCodeExtract, sport="Soccer"))
    eventThreeDisplayButton.place(x=330, y=200, width=140, height=150)
    eventFourDisplayButton = Button(page2, text='Sport\nFour', font=('yu gothic ui', 30, 'bold'), bg='#556d7c',
                                    fg='black', command=lambda: openInfoSport('4', 'Sport', eventFourCodeExtract,
                                                                              sport="Volleyball"))
    eventFourDisplayButton.place(x=480, y=200, width=140, height=150)
    eventFiveDisplayButton = Button(page2, text='Sport\nFive', font=('yu gothic ui', 30, 'bold'), bg='#556d7c',
                                    fg='black',
                                    command=lambda: openInfoSport('5', 'Sport', eventFiveCodeExtract, sport="Chess"))
    eventFiveDisplayButton.place(x=630, y=200, width=140, height=150)
    eventSixDisplayButton = Button(page2, text='Fun\nOne', font=('yu gothic ui', 30, 'bold'), bg='#556d7c',
                                   fg='black',
                                   command=lambda: openInfoSport('1', 'Fun', eventSixCodeExtract, sport="Prom"))
    eventSixDisplayButton.place(x=30, y=400, width=140, height=150)
    eventSevenDisplayButton = Button(page2, text='Fun\nTwo', font=('yu gothic ui', 30, 'bold'), bg='#556d7c',
                                     fg='black', command=lambda: openInfoSport('2', 'Fun', eventSevenCodeExtract,
                                                                               sport="Spirit Night"))
    eventSevenDisplayButton.place(x=180, y=400, width=140, height=150)
    eventEightDisplayButton = Button(page2, text='Fun\nThree', font=('yu gothic ui', 30, 'bold'), bg='#556d7c',
                                     fg='black', command=lambda: openInfoSport('3', 'Fun', eventEightCodeExtract,
                                                                               sport="Homecoming"))
    eventEightDisplayButton.place(x=330, y=400, width=140, height=150)
    eventNineDisplayButton = Button(page2, text='Fun\nFour', font=('yu gothic ui', 30, 'bold'), bg='#556d7c',
                                    fg='black', command=lambda: openInfoSport('4', 'Fun', eventNineCodeExtract,
                                                                              sport="International Night"))
    eventNineDisplayButton.place(x=480, y=400, width=140, height=150)
    eventTenDisplayButton = Button(page2, text='Fun\nFive', font=('yu gothic ui', 30, 'bold'), bg='#556d7c',
                                   fg='black',
                                   command=lambda: openInfoSport('5', 'Fun', eventTenCodeExtract, sport="Drama"))
    eventTenDisplayButton.place(x=630, y=400, width=140, height=150)

    # place the buttons

    homeButton.place(x=0, y=0, width=200)
    eventButton.place(x=200, y=0, width=200)
    attendanceButton.place(x=400, y=0, width=200)
    lotteryButton.place(x=600, y=0, width=200)

    # ======== Page 3 ===========
    homeButton = Button(page3, text='Home', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page1),
                        bg='#36454F', fg='white')
    eventButton = Button(page3, text='Events', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page2),
                         bg='#36454F', fg='white')
    attendanceButton = Button(page3, text='Participation', font=('yu gothic ui', 15, 'bold'),
                              command=lambda: show_frame(page3), bg='#36454F', fg='white')
    lotteryButton = Button(page3, text='Prizes', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page4),
                           bg='#36454F', fg='white')

    # place the buttons

    homeButton.place(x=0, y=0, width=200)
    eventButton.place(x=200, y=0, width=200)
    attendanceButton.place(x=400, y=0, width=200)
    lotteryButton.place(x=600, y=0, width=200)

    # ======== Page 4-Prizes ===========
    homeButton = Button(page4, text='Home', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page1),
                        bg='#36454F', fg='white')
    eventButton = Button(page4, text='Events', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page2),
                         bg='#36454F', fg='white')
    attendanceButton = Button(page4, text='Participation', font=('yu gothic ui', 15, 'bold'),
                              command=lambda: show_frame(page3), bg='#36454F', fg='white')
    lotteryButton = Button(page4, text='Prizes', font=('yu gothic ui', 15, 'bold'), command=lambda: show_frame(page4),
                           bg='#36454F', fg='white')

    scoreBoardLabel = Label(page4, text='Scoreboard:', font=('yu gothic ui', 30, 'bold'), bg='#36454F', fg='white')
    scoreBoardLabel.place(x=30, y=100)

    yourLabel = Label(page4, text='Your Score:', font=('yu gothic ui', 15, 'bold'), bg='#36454F', fg='white')
    yourLabel.place(x=30, y=180)

    leaderLabel = Label(page4, text="Leader's Score:", font=('yu gothic ui', 15, 'bold'), bg='#36454F', fg='white')
    leaderLabel.place(x=30, y=230)

    df = pd.read_excel('Book1.xlsx')
    maxCol = len(df['First Name'])
    yourPosition(index, page4, x=140, y=180, tot=maxCol, x2=170, y2=230, x3=30, y3=340, x4=30, y4=460)

    leaderNameLabel = Label(page4, text='Leader: ', font=('yu gothic ui', 20, 'bold'), bg='#36454F', fg='white')
    leaderNameLabel.place(x=30, y=300)

    showPrize = Label(page4, text="Your Prize:  ", font=('yu gothic ui', 16, 'bold'), bg='#36454F', fg='white')
    showPrize.place(x=30, y=420)

    # place the buttons

    homeButton.place(x=0, y=0, width=200)
    eventButton.place(x=200, y=0, width=200)
    attendanceButton.place(x=400, y=0, width=200)
    lotteryButton.place(x=600, y=0, width=200)

    window.mainloop()


def getCredentials(entry, entry2, window):
    credential1 = entry.get()
    credential2 = entry2.get()
    # Open Excel File

    df = pd.read_excel('Book1.xlsx', sheet_name=0)
    firstNameList = df['First Name'].to_list()
    lastNameList = df['Last Name'].to_list()
    schoolNameList = df['School Name'].to_list()
    schoolYearList = df['School Year'].to_list()
    usernameList = df['Email'].to_list()
    passwordList = df['Password'].to_list()

    # Cross Check Values

    if credential1 in usernameList:
        ind = usernameList.index(credential1)

        if passwordList[ind] == credential2:
            if usernameList[ind] == "admin@gmail.com" and passwordList[ind] == "admin":
                window.destroy()
                MainEventSupervisor()
                print("Admin Success")
            else:
                print("Success")
                window.destroy()
                dashboard(firstNameList[ind], lastNameList[ind], schoolNameList[ind], str(schoolYearList[ind]),
                          usernameList[ind], (ind + 2))
        else:
            print("Invalid Combination")
            tkinter.messagebox.showerror("Error", "Invalid Combination")
            window.focus_force()

    else:
        print("Invalid Combination")
        tkinter.messagebox.showerror("Error", "Invalid Combination")
        window.focus_force()


def getRegisterCredentials(entry1, entry2, entry3, entry4, entry5, entry6, window):
    firstName = entry1.get()
    lastName = entry2.get()
    schoolName = entry3.get()
    schoolYear = entry4.get()
    username = entry5.get()
    password = entry6.get()

    # Open Excel sheet through OpenPyXl first
    wb = op.load_workbook('Book1.xlsx')
    ws = wb.active

    # Open Excel sheet through Pandas for crosschecking repetition of usernames
    wb2 = pd.read_excel('Book1.xlsx', sheet_name=0)
    emailList = wb2['Email'].to_list()
    print(emailList)
    match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', username)
    if match is None:
        tkinter.messagebox.showerror("Error", "Enter a valid Email ID!")

    if username in emailList:
        print("Email already registered!")
        tkinter.messagebox.showerror("Error", "Email already registered!")
        window.focus_force()
    else:
        ws.append([firstName, lastName, schoolName, schoolYear, username, password])
        wb.save('Book1.xlsx')
        window.destroy()


def registerNames():
    registerWindow = Tk()
    registerWindow.geometry('700x600')
    registerWindow.title('Create Account')
    registerWindow.config(bg="#36454F")
    registerWindow.focus_force()  # Focus on the register window

    # Title
    titleLabel = Label(registerWindow, text="Account Creation", font=("yu gothic ui", 25, "bold"), fg="white",
                       bg="#36454F")
    titleLabel.place(x=20, y=30)

    # Info Entry
    # First Name
    firstNameLabel = Label(registerWindow, text="First Name", font=("yu gothic ui", 15, "bold"), fg="#4f4e4d",
                           bg="#36454F")
    firstNameLabel.place(x=20, y=100)
    firstNameEntry = Entry(registerWindow, highlightthickness=0, relief=FLAT, bg="#36454F", fg="#6b6a69",
                           font=("yu gothic ui", 15, "bold"))
    firstNameEntry.place(x=25, y=140, width=270)
    firstNameLine = Canvas(registerWindow, width=270, height=2.0, bg="#bdb9b1", highlightthickness=0)
    firstNameLine.place(x=25, y=170)

    # Last Name
    lastNameLabel = Label(registerWindow, text="Last Name", font=("yu gothic ui", 15, "bold"), fg="#4f4e4d",
                          bg="#36454F")
    lastNameLabel.place(x=20, y=190)
    lastNameEntry = Entry(registerWindow, highlightthickness=0, relief=FLAT, bg="#36454F", fg="#6b6a69",
                          font=("yu gothic ui", 15, "bold"))
    lastNameEntry.place(x=25, y=230, width=270)
    lastNameLine = Canvas(registerWindow, width=270, height=2.0, bg="#bdb9b1", highlightthickness=0)
    lastNameLine.place(x=25, y=260)

    # School
    schoolLabel = Label(registerWindow, text="School Name (Abbreviated)", font=("yu gothic ui", 15, "bold"),
                        fg="#4f4e4d",
                        bg="#36454F")
    schoolLabel.place(x=20, y=280)
    schoolEntry = Entry(registerWindow, highlightthickness=0, relief=FLAT, bg="#36454F", fg="#6b6a69",
                        font=("yu gothic ui", 15, "bold"))
    schoolEntry.place(x=25, y=320, width=270)
    schoolLine = Canvas(registerWindow, width=270, height=2.0, bg="#bdb9b1", highlightthickness=0)
    schoolLine.place(x=25, y=350)

    # School Year
    schoolYearLabel = Label(registerWindow, text="School Year (in numbers)", font=("yu gothic ui", 15, "bold"),
                            fg="#4f4e4d",
                            bg="#36454F")
    schoolYearLabel.place(x=20, y=370)
    schoolYearEntry = Entry(registerWindow, highlightthickness=0, relief=FLAT, bg="#36454F", fg="#6b6a69",
                            font=("yu gothic ui", 15, "bold"))
    schoolYearEntry.place(x=25, y=410, width=270)
    schoolYearLine = Canvas(registerWindow, width=270, height=2.0, bg="#bdb9b1", highlightthickness=0)
    schoolYearLine.place(x=25, y=440)

    # Username
    newUsernameLabel = Label(registerWindow, text="Email ID", font=("yu gothic ui", 15, "bold"),
                             fg="#4f4e4d",
                             bg="#36454F")
    newUsernameLabel.place(x=320, y=190)
    newUsernameEntry = Entry(registerWindow, highlightthickness=0, relief=FLAT, bg="#36454F", fg="#6b6a69",
                             font=("yu gothic ui", 15, "bold"))
    newUsernameEntry.place(x=325, y=230, width=270)
    newUsernameLine = Canvas(registerWindow, width=270, height=2.0, bg="#bdb9b1", highlightthickness=0)
    newUsernameLine.place(x=325, y=260)

    # Password
    newPasswordLabel = Label(registerWindow, text="New Password", font=("yu gothic ui", 15, "bold"),
                             fg="#4f4e4d",
                             bg="#36454F")
    newPasswordLabel.place(x=320, y=280)
    newPasswordEntry = Entry(registerWindow, highlightthickness=0, relief=FLAT, bg="#36454F", fg="#6b6a69",
                             font=("yu gothic ui", 15, "bold"))
    newPasswordEntry.place(x=325, y=320, width=270)
    newPasswordLine = Canvas(registerWindow, width=270, height=2.0, bg="#bdb9b1", highlightthickness=0)
    newPasswordLine.place(x=325, y=350)

    # Button
    registerButton = Button(registerWindow, text="Create Account", font=("yu gothic ui", 15, "bold"), bg="#36454F",
                            fg="white", borderwidth=0, command=
                            lambda: getRegisterCredentials(firstNameEntry, lastNameEntry, schoolEntry, schoolYearEntry,
                                                           newUsernameEntry, newPasswordEntry, registerWindow))
    registerButton.place(x=20, y=450)

    registerWindow.mainloop()


def page():
    window = Tk()
    window.geometry("700x600")
    window.config(bg="#36454F")
    window.title('Log-in')

    loginText = Label(window, text="User Log-in", font=('yu gothic ui', 30, 'bold'), fg='#bdb9b1', bg="#36454F")
    loginText.place(x=20, y=20)

    username = Label(window, text="Email ID", bg="#36454F", fg="#bdb9b1", font=("yu gothic ui", 15, "bold"))
    username.place(x=20, y=110)
    usernameEntry = Entry(window, highlightthickness=0, relief=FLAT, bg="#36454F", fg="#bdb9b1",
                          font=("yu gothic ui", 15, "bold"))
    usernameEntry.place(x=25, y=150, width=270)
    usernameLine = Canvas(window, width=270, height=2.0, bg="#bdb9b1", highlightthickness=0)
    usernameLine.place(x=25, y=190)

    password = Label(window, text="Password", bg="#36454F", fg="#bdb9b1", font=("yu gothic ui", 15, "bold"))
    password.place(x=20, y=210)
    passwordEntry = Entry(window, highlightthickness=0, relief=FLAT, bg="#36454F", fg="#bdb9b1",
                          font=("yu gothic ui", 15, "bold"))
    passwordEntry.place(x=25, y=250, width=270)
    passwordLine = Canvas(window, width=270, height=2.0, bg="#bdb9b1", highlightthickness=0)
    passwordLine.place(x=25, y=290)

    loginButton = Button(window, text="Enter", font=('yu gothic ui', 15, 'bold'), fg="white", bg="#36454F",
                         borderwidth=0, command=lambda: getCredentials(usernameEntry, passwordEntry, window))
    loginButton.place(x=20, y=320)

    registerButton = Button(window, text="Do not have an account? Click to Register", font=('yu gothic ui', 15, 'bold'),
                            fg="white", bg="#36454F",
                            borderwidth=0, command=lambda: registerNames())
    registerButton.place(x=20, y=360)

    window.mainloop()


if __name__ == '__main__':
    page()
